#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from __future__ import annotations

import math
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd

from PySide6.QtCore import Qt
from PySide6.QtGui import QGuiApplication
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QDoubleSpinBox, QTextEdit, QMessageBox,
    QComboBox
)

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


# =========================
# Lens tables
# =========================

LENSES_RZ34_RQ35 = {
    "DLE95": (0.4, 0.4),
    "LE50": (0.7, 0.7),
    "LE6": (0.9, 1.1),
    "LE10": (1.3, 1.7),
    "LE20": (1.7, 2.4),
    "LE30": (2.3, 4.7),
    "LE40": (4.6, 7.4),
    "LE8": (7.3, 13.8),
}
LENSES_RQ25 = {
    "DLE95": (0.44, 0.44),
    "LE50": (0.84, 0.84),
    "LE6": (1.11, 1.32),
    "LE10": (1.56, 2.01),
    "LE20": (2.0, 2.9),
    "LE30": (2.89, 5.6),
    "LE40": (5.54, 8.9),
    "LE8": (8.83, 16.6),
}


# =========================
# Helpers
# =========================

class DSU:
    def __init__(self, n: int):
        self.p = list(range(n))
        self.r = [0] * n

    def find(self, a: int) -> int:
        while self.p[a] != a:
            self.p[a] = self.p[self.p[a]]
            a = self.p[a]
        return a

    def union(self, a: int, b: int) -> None:
        ra, rb = self.find(a), self.find(b)
        if ra == rb:
            return
        if self.r[ra] < self.r[rb]:
            ra, rb = rb, ra
        self.p[rb] = ra
        if self.r[ra] == self.r[rb]:
            self.r[ra] += 1


def fmt_int(n: int) -> str:
    return f"{n:,}"


def safe_float(v) -> Optional[float]:
    try:
        if pd.isna(v):
            return None
        return float(v)
    except Exception:
        return None


def projector_model(resx, resy, lumens) -> Optional[str]:
    """
    Map resolution+lumens to projector model per rules:
    1920x1200 + 31000 => Panasonic PT-RZ34k
    3840x2400 + 31000 => Panasonic PT-RQ35k
    3840x2400 + 21000 => Panasonic PT-RQ25k
    """
    try:
        rx, ry = int(resx), int(resy)
    except Exception:
        return None

    lum = None
    if not pd.isna(lumens):
        try:
            lum = int(float(lumens))
        except Exception:
            lum = None

    if rx == 1920 and ry == 1200 and lum == 31000:
        return "Panasonic PT-RZ34k"
    if rx == 3840 and ry == 2400 and lum == 31000:
        return "Panasonic PT-RQ35k"
    if rx == 3840 and ry == 2400 and lum == 21000:
        return "Panasonic PT-RQ25k"

    # fallback by resolution if lumens missing
    if rx == 1920 and ry == 1200:
        return "Panasonic PT-RZ34k"
    if rx == 3840 and ry == 2400:
        return "Panasonic PT-RQ35k"
    return None


def choose_lens(model: str, throw_ratio) -> Optional[str]:
    """
    Choose lens by throw ratio ranges.
    Overlap rule: if overlapping, pick lens with smaller range width.
    """
    if pd.isna(throw_ratio):
        return None
    r = float(throw_ratio)
    lenses = LENSES_RQ25 if model == "Panasonic PT-RQ25k" else LENSES_RZ34_RQ35

    candidates: List[Tuple[str, float, float]] = []
    for name, (a, b) in lenses.items():
        lo, hi = (a, b) if a <= b else (b, a)
        if lo <= r <= hi:
            candidates.append((name, lo, hi))

    if not candidates:
        return None

    candidates.sort(key=lambda x: ((x[2] - x[1]), x[2], x[0]))
    return candidates[0][0]


# =========================
# Tower rule selection
# =========================

class TowerRule:
    OR_AXIS = "OR (dx < D OR dy < D)"
    AND_AXIS = "AND (dx < D AND dy < D)"
    EUCLIDEAN = "EUCLIDEAN (sqrt(dx^2+dy^2) < D)"


def same_tower(dx: float, dy: float, D: float, rule: str) -> bool:
    if rule == TowerRule.OR_AXIS:
        return (dx < D) or (dy < D)
    if rule == TowerRule.AND_AXIS:
        return (dx < D) and (dy < D)
    if rule == TowerRule.EUCLIDEAN:
        return (dx * dx + dy * dy) ** 0.5 < D
    return (dx < D) or (dy < D)


@dataclass
class ParseResult:
    report: Dict
    details_df: pd.DataFrame
    warnings: List[str]


def compute_report(df: pd.DataFrame, tower_distance_m: float, tower_rule: str) -> ParseResult:
    # Expected columns for the example CSV format.
    col_resx = "Projector_Native-Rez-X"
    col_resy = "Projector_Native-Rez-Y"
    col_lum  = "Projector_Lumens(lux)"
    col_stack= "Projector_Qte(Stack)"
    col_tr   = "Projector_Trow-Ratio"
    col_x, col_y, col_z = "Lens_X", "Lens_Y", "Lens_Z"
    col_ill  = "Target_Illuminance"

    required = [col_resx, col_resy, col_lum, col_stack, col_tr, col_x, col_y, col_z, col_ill]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"CSV is missing required columns: {missing}")

    df = df.copy()
    warnings: List[str] = []

    df["Model"] = [projector_model(a, b, c) for a, b, c in zip(df[col_resx], df[col_resy], df[col_lum])]
    df["StackSize"] = pd.to_numeric(df[col_stack], errors="coerce").fillna(1).astype(int)
    df["ThrowRatio"] = pd.to_numeric(df[col_tr], errors="coerce")
    df["LensModel"] = pd.Series([None] * len(df), index=df.index, dtype="object")

    for idx, (m, tr) in enumerate(zip(df["Model"], df["ThrowRatio"]), start=1):
        if not m:
            warnings.append(f"Row {idx}: unknown projector model (resolution/lumens mapping failed).")
            continue
        lens = choose_lens(m, tr)
        df.loc[df.index[idx - 1], "LensModel"] = lens
        if lens is None:
            trf = safe_float(tr)
            if trf is None:
                warnings.append(f"Row {idx}: throw ratio is empty/invalid -> cannot choose lens.")
            else:
                warnings.append(f"Row {idx}: throw ratio {trf:g} does not match any lens for {m}.")

    video_signals = int(len(df))
    total_projectors = int(df["StackSize"].sum())
    stacks_by_size = df["StackSize"].value_counts().sort_index()
    proj_by_model = df.groupby("Model")["StackSize"].sum().sort_values(ascending=False)
    lens_counts = df.groupby("LensModel")["StackSize"].sum().sort_values(ascending=False)

    median_illum = float(pd.to_numeric(df[col_ill], errors="coerce").median())
    matrices_48 = int(math.ceil(total_projectors / 48.0))

    is_4k = (df[col_resx].astype(int) == 3840) & (df[col_resy].astype(int) == 2400)
    signals_4k = int(is_4k.sum())
    signals_lt4k = int((~is_4k).sum())

    actors_hd = int(math.ceil(signals_lt4k / 16.0)) if signals_lt4k > 0 else 0
    actors_4k = int(math.ceil(signals_4k / 4.0)) if signals_4k > 0 else 0
    director = 1
    understudy = int(math.ceil((actors_hd + actors_4k + director) / 3.0))

    # Towers
    height = pd.to_numeric(df[col_z], errors="coerce")
    df["OnTower"] = (height > 3.0).fillna(False)

    tower_df = df[df["OnTower"]].reset_index(drop=True)
    tower_count = 0
    if len(tower_df) > 0:
        xs = pd.to_numeric(tower_df[col_x], errors="coerce").to_numpy()
        ys = pd.to_numeric(tower_df[col_y], errors="coerce").to_numpy()
        n = len(tower_df)
        dsu = DSU(n)
        for i in range(n):
            for j in range(i + 1, n):
                dx = abs(xs[i] - xs[j])
                dy = abs(ys[i] - ys[j])
                if same_tower(dx, dy, tower_distance_m, tower_rule):
                    dsu.union(i, j)
        tower_count = len({dsu.find(i) for i in range(n)})

    report = {
        "Projectors by model": proj_by_model.to_dict(),
        "Total projectors": total_projectors,
        "Stacks (count) by size": stacks_by_size.to_dict(),
        "Lenses (qty) by model": lens_counts.to_dict(),
        "Video signals": video_signals,
        "Median Target_Illuminance": median_illum,
        "disguise vx4+ director": director,
        "disguise vx4+ actors (QUAD-DVI, <=WUXGA)": actors_hd,
        "disguise vx4+ actors (HDMI, 4K)": actors_4k,
        "disguise vx4+ understudy": understudy,
        "HDMI matrices 48x48": matrices_48,
        "Towers": int(tower_count),
        "Tower distance (m)": float(tower_distance_m),
        "Tower rule": tower_rule,
    }
    return ParseResult(report=report, details_df=df, warnings=warnings)


def build_report_text(result: ParseResult, title: str) -> str:
    r = result.report
    lines: List[str] = []
    lines.append(f"File: {title}")
    lines.append("")
    lines.append("Projectors")
    for model, qty in r["Projectors by model"].items():
        lines.append(f"  - {model}: {fmt_int(int(qty))}")
    lines.append(f"  - Total projectors: {fmt_int(int(r['Total projectors']))}")
    lines.append("")
    lines.append("Stacks (positions)")
    for size, count in r["Stacks (count) by size"].items():
        lines.append(f"  - Stack size {size}: {fmt_int(int(count))}")
    lines.append("")
    lines.append("Lenses")
    if len(r["Lenses (qty) by model"]) == 0:
        lines.append("  - (no data)")
    else:
        for lens, qty in r["Lenses (qty) by model"].items():
            if lens is None or str(lens) == "nan":
                continue
            lines.append(f"  - {lens}: {fmt_int(int(qty))}")
    lines.append("")
    lines.append("Video signals")
    lines.append(f"  - Count: {fmt_int(int(r['Video signals']))}")
    lines.append("")
    lines.append("Target Illuminance")
    lines.append(f"  - Median: {r['Median Target_Illuminance']:.2f}")
    lines.append("")
    lines.append("disguise vx4+")
    lines.append(f"  - Dedicated director: {fmt_int(int(r['disguise vx4+ director']))}")
    lines.append(f"  - Actor (QUAD-DVI, <=WUXGA): {fmt_int(int(r['disguise vx4+ actors (QUAD-DVI, <=WUXGA)']))}")
    lines.append(f"  - Actor (HDMI, 4K): {fmt_int(int(r['disguise vx4+ actors (HDMI, 4K)']))}")
    lines.append(f"  - Understudy: {fmt_int(int(r['disguise vx4+ understudy']))}")
    lines.append("")
    lines.append("HDMI matrices 48x48")
    lines.append(f"  - Count: {fmt_int(int(r['HDMI matrices 48x48']))}")
    lines.append("")
    lines.append("Towers")
    lines.append(f"  - Rule: {r['Tower rule']}")
    lines.append(f"  - Threshold (m): {r['Tower distance (m)']:.2f}")
    lines.append(f"  - Tower count: {fmt_int(int(r['Towers']))}")

    if result.warnings:
        lines.append("")
        lines.append("Warnings")
        for w in result.warnings[:30]:
            lines.append(f"  - {w}")
        if len(result.warnings) > 30:
            lines.append(f"  - ... and {len(result.warnings) - 30} more warnings")

    return "\n".join(lines)


def write_excel(result: ParseResult, out_path: str) -> None:
    r = result.report
    df_details = result.details_df

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Equipment summary"
    ws["A1"].font = Font(bold=True, size=14)

    row = 3
    def kv(k, v):
        nonlocal row
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    kv("Total projectors", r["Total projectors"])
    kv("Video signals", r["Video signals"])
    kv("Median Target_Illuminance", r["Median Target_Illuminance"])
    kv("HDMI matrices 48x48", r["HDMI matrices 48x48"])
    kv("Tower rule", r["Tower rule"])
    kv("Tower distance (m)", r["Tower distance (m)"])
    kv("Towers", r["Towers"])
    row += 1

    ws[f"A{row}"] = "Projectors by model"; ws[f"A{row}"].font = Font(bold=True); row += 1
    for k, v in r["Projectors by model"].items():
        kv(k, v)
    row += 1

    ws[f"A{row}"] = "Stacks by size"; ws[f"A{row}"].font = Font(bold=True); row += 1
    for k, v in r["Stacks (count) by size"].items():
        kv(f"{k}-stack positions", v)
    row += 1

    ws[f"A{row}"] = "Lenses by model (qty)"; ws[f"A{row}"].font = Font(bold=True); row += 1
    for k, v in r["Lenses (qty) by model"].items():
        kv(k, v)

    if result.warnings:
        row += 1
        ws[f"A{row}"] = "Warnings"; ws[f"A{row}"].font = Font(bold=True); row += 1
        for w in result.warnings:
            ws[f"A{row}"] = w
            row += 1

    ws2 = wb.create_sheet("Details")
    for rr in dataframe_to_rows(df_details, index=False, header=True):
        ws2.append(rr)

    nrows, ncols = df_details.shape
    last_col = get_column_letter(ncols)
    tbl = Table(displayName="Projectors", ref=f"A1:{last_col}{nrows+1}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws2.add_table(tbl)
    ws2.freeze_panes = "A2"

    # autosize
    for wsx in (ws, ws2):
        for col in wsx.columns:
            maxlen = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is None:
                    continue
                maxlen = max(maxlen, len(str(cell.value)))
            wsx.column_dimensions[col_letter].width = min(max(10, maxlen + 2), 60)

    wb.save(out_path)


# =========================
# GUI app
# =========================

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Projector CSV -> Equipment List")
        self.resize(980, 700)

        self.csv_path: Optional[str] = None
        self.last_result: Optional[ParseResult] = None

        root = QWidget()
        self.setCentralWidget(root)
        v = QVBoxLayout(root)

        top = QHBoxLayout()
        v.addLayout(top)

        self.btn_open = QPushButton("Open CSV...")
        self.btn_open.clicked.connect(self.open_csv)
        top.addWidget(self.btn_open)

        top.addWidget(QLabel("Tower threshold (m):"))
        self.spin_dist = QDoubleSpinBox()
        self.spin_dist.setRange(0.1, 1000.0)
        self.spin_dist.setDecimals(2)
        self.spin_dist.setSingleStep(0.5)
        self.spin_dist.setValue(4.0)
        self.spin_dist.valueChanged.connect(self.recompute_if_possible)
        top.addWidget(self.spin_dist)

        top.addWidget(QLabel("Tower rule:"))
        self.combo_rule = QComboBox()
        self.combo_rule.addItems([TowerRule.OR_AXIS, TowerRule.AND_AXIS, TowerRule.EUCLIDEAN])
        self.combo_rule.currentIndexChanged.connect(self.recompute_if_possible)
        top.addWidget(self.combo_rule)

        self.btn_copy = QPushButton("Copy report")
        self.btn_copy.clicked.connect(self.copy_report)
        self.btn_copy.setEnabled(False)
        top.addWidget(self.btn_copy)

        self.btn_save = QPushButton("Save Excel...")
        self.btn_save.clicked.connect(self.save_excel)
        self.btn_save.setEnabled(False)
        top.addWidget(self.btn_save)

        top.addStretch(1)

        self.lbl_file = QLabel("No file selected")
        self.lbl_file.setTextInteractionFlags(Qt.TextSelectableByMouse)
        v.addWidget(self.lbl_file)

        self.text = QTextEdit()
        self.text.setReadOnly(True)
        self.text.setPlaceholderText("Open a CSV file to see the report here.")
        v.addWidget(self.text, 1)

    def open_csv(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select CSV file",
            "",
            "CSV files (*.csv);;All files (*.*)"
        )
        if not path:
            return
        self.csv_path = path
        self.lbl_file.setText(path)
        self.compute()

    def recompute_if_possible(self):
        if self.csv_path:
            self.compute()

    def compute(self):
        try:
            df = pd.read_csv(self.csv_path)
            dist = float(self.spin_dist.value())
            rule = self.combo_rule.currentText()
            result = compute_report(df, dist, rule)

            self.last_result = result
            report = build_report_text(result, title=self.csv_path.split("/")[-1])
            self.text.setPlainText(report)

            self.btn_save.setEnabled(True)
            self.btn_copy.setEnabled(True)
        except Exception as e:
            self.last_result = None
            self.btn_save.setEnabled(False)
            self.btn_copy.setEnabled(False)
            QMessageBox.critical(self, "Error", str(e))

    def copy_report(self):
        s = self.text.toPlainText()
        if not s.strip():
            return
        QGuiApplication.clipboard().setText(s)
        QMessageBox.information(self, "Copied", "Report copied to clipboard.")

    def save_excel(self):
        if not self.last_result:
            return
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel report",
            "equipment_report.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            write_excel(self.last_result, path)
            QMessageBox.information(self, "Saved", f"Saved:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


def main():
    app = QApplication([])
    w = MainWindow()
    w.show()
    app.exec()


if __name__ == "__main__":
    main()
